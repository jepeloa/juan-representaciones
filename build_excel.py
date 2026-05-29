#!/usr/bin/env python3
"""Consolidate product price lists from zip into a single Excel.

Outputs /home/javier/juan/Catalogo_Consolidado.xlsx with:
 - "Consolidado": all products (flat)
 - One sheet per supplier
 - "Para revisar": files that couldn't be parsed (image-only PDFs / catalog-only PDFs)
"""
from __future__ import annotations
import re
import sys
import traceback
from pathlib import Path
from dataclasses import dataclass, field
from typing import Iterator

import openpyxl
import pdfplumber
import pandas as pd

BASE = Path('/home/javier/juan/Listas actualizadas')
TXT = BASE / '_txt'
TXT_OCR = BASE / '_txt_ocr'
OUT = Path('/home/javier/juan/Catalogo_Consolidado.xlsx')

COLUMNS = [
    'Proveedor', 'Categoria', 'Codigo', 'Producto', 'Descripcion',
    'Precio', 'Moneda', 'IVA', 'Cantidad x Bulto', 'Codigo Barras',
    'Observaciones', 'Archivo origen', 'Pagina PDF',
]


def to_money(v):
    """Normalize a price cell to float, or None.

    Handles both Argentine ('1.234.567,89') and English ('1,234,567.89') number
    formats, plus single-thousands ('120.000' AR or '120,000' EN) by checking
    whether every group after the first separator has exactly 3 digits.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Remove currency markers
    s = re.sub(r'(?i)u\$?s?d?|usd|\$|AR\$', '', s).strip()
    s = s.replace(' ', '')
    if not s:
        return None
    # Both separators present: the last one wins as decimal.
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        parts = s.split(',')
        # Comma as thousands when every segment after the first has 3 digits AND
        # there are 2+ groups (e.g. '120,000' or '1,234,567'). Otherwise decimal.
        if len(parts) >= 2 and all(len(p) == 3 and p.isdigit() for p in parts[1:]):
            s = s.replace(',', '')
        elif len(parts) == 2 and len(parts[1]) in (1, 2):
            s = parts[0].replace('.', '') + '.' + parts[1]
        else:
            s = s.replace(',', '')
    elif '.' in s:
        parts = s.split('.')
        # Dot as thousands when every segment after the first has 3 digits
        # (e.g. '120.000' = 120000, '1.234.567' = 1234567). Otherwise decimal.
        if len(parts) >= 2 and all(len(p) == 3 and p.isdigit() for p in parts[1:]):
            s = ''.join(parts)
    try:
        return float(s)
    except ValueError:
        return None


def detect_currency(txt) -> str:
    if txt is None:
        return 'ARS'
    s = str(txt).upper()
    if 'U$S' in s or 'U$D' in s or 'USD' in s or 'U$' in s:
        return 'USD'
    return 'ARS'


def row(proveedor, **kw):
    r = {c: None for c in COLUMNS}
    r['Proveedor'] = proveedor
    r.update({k: v for k, v in kw.items() if k in r})
    return r


def lines_with_pages(txt_path: Path) -> list[tuple[int, str]]:
    """Yield (page_number, line) tuples from a text file.

    Recognises two page-separator formats:
      - '\\f' (form-feed) — output of `pdftotext -layout`
      - '===== PAGE N =====' — output of our OCR pipeline (ocr_pending.py)
    """
    raw = txt_path.read_text(encoding='utf-8', errors='replace')
    out: list[tuple[int, str]] = []

    if '===== PAGE' in raw:
        page = 1
        for ln in raw.splitlines():
            m = re.match(r'^=+\s*PAGE\s+(\d+)\s*=+\s*$', ln)
            if m:
                page = int(m.group(1))
                continue
            out.append((page, ln))
        return out

    # pdftotext layout: form-feed separates pages
    page = 1
    for chunk in raw.split('\f'):
        for ln in chunk.splitlines():
            out.append((page, ln))
        page += 1
    return out


# ============ DUNIA XLSX ============
def parse_dunia_xlsx() -> Iterator[dict]:
    """Dunia consolidated XLSX (mayo 2026).

    Layout (single sheet 'Hoja1'):
        row 1: leyenda stock
        row 2: headers (IMAGEN | CÓDIGO INTERNO | NOMBRE DE ARTÍCULO | PRECIOS MARZO | CANTIDAD POR BULTO CERRADO | CÓDIGOS | CÓDIGO DE BARRAS)
        row 3+: category rows have only col A populated, then data rows.
    """
    fname = 'DUNIA-MAQUINARIAS - ABRIL.xlsx'
    path = BASE / 'Dunia' / fname
    if not path.exists():
        return
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None
    current_cat = 'Maquinaria'
    for i, vals in enumerate(ws.iter_rows(values_only=True)):
        vals = list(vals)
        if headers is None:
            up = ' '.join(str(v or '') for v in vals).upper()
            if 'NOMBRE' in up and ('PRECIO' in up or 'PRECIOS' in up):
                headers = [str(v or '').strip().upper() for v in vals]
            continue

        def col(name):
            for hi, h in enumerate(headers):
                if name in h:
                    return vals[hi] if hi < len(vals) else None
            return None

        codigo = col('CÓDIGO INTERNO') or col('CODIGO INTERNO')
        nombre = col('NOMBRE')
        precio = col('PRECIO')
        cantidad = col('CANTIDAD')
        barras = col('CÓDIGO DE BARRAS') or col('CODIGO DE BARRAS')
        extra_cod = col('CÓDIGOS') or col('CODIGOS')

        # Category-only row: col B (CÓDIGO INTERNO) is empty but col A has text
        if not codigo and not nombre and not precio and vals[0]:
            current_cat = str(vals[0]).strip()
            continue
        if not nombre:
            continue
        price_num = to_money(precio)
        obs = str(precio) if (price_num is None and precio not in (None, '')) else None
        extra_code_str = str(extra_cod).strip() if extra_cod not in (None, '') else None
        yield row(
            'Dunia',
            Categoria=current_cat,
            Codigo=str(codigo).strip() if codigo else None,
            Producto=str(nombre).strip(),
            Descripcion=None,
            Precio=price_num,
            Moneda='ARS',
            IVA=None,
            **{'Cantidad x Bulto': cantidad if isinstance(cantidad, (int, float)) else None,
               'Codigo Barras': str(barras).strip() if barras else None},
            Observaciones=' | '.join(x for x in [
                f'cod alt:{extra_code_str}' if extra_code_str else None,
                obs,
            ] if x) or None,
            **{'Archivo origen': fname},
        )
    wb.close()


# ============ HAVARD Tramontina (clean table) ============
def parse_havard_tramontina() -> Iterator[dict]:
    path = BASE / 'Havard' / 'LISTA MAYORISTA TRAMONTINA HAVARD.pdf'
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            pdf_page = pi + 1
            for tbl in page.extract_tables():
                if not tbl or len(tbl[0]) < 5:
                    continue
                header = [str(c or '').upper() for c in tbl[0]]
                if 'CODIGO' not in header[0] and 'CÓDIGO' not in header[0]:
                    continue
                for r in tbl[1:]:
                    if not r or not r[0]:
                        continue
                    codigo = (r[0] or '').strip()
                    articulo = (r[2] or '').strip() if len(r) > 2 else ''
                    if not articulo:
                        continue
                    iva = (r[3] or '').strip() if len(r) > 3 else ''
                    precio_raw = (r[4] or '').strip() if len(r) > 4 else ''
                    qty = (r[5] or '').strip() if len(r) > 5 else ''
                    stock = (r[6] or '').strip() if len(r) > 6 else ''
                    # "17 USD" -> price 17, currency USD
                    m = re.match(r'([\d.,]+)\s*(USD|U\$S|U\$D|\$)?', precio_raw)
                    price_num = to_money(m.group(1)) if m else to_money(precio_raw)
                    currency = 'USD' if 'USD' in precio_raw.upper() or 'U$' in precio_raw.upper() else 'ARS'
                    # articulo puede tener descripcion multi-linea
                    parts = articulo.split('\n')
                    producto = parts[0].strip()
                    descripcion = '\n'.join(p.strip() for p in parts[1:]) if len(parts) > 1 else None
                    yield row(
                        'Havard',
                        Categoria='Tramontina',
                        Codigo=codigo,
                        Producto=producto,
                        Descripcion=descripcion,
                        Precio=price_num,
                        Moneda=currency,
                        IVA=iva,
                        Observaciones=f'stock:{stock}; qty:{qty}' if (stock or qty) else None,
                        **{'Archivo origen': path.name, 'Pagina PDF': pdf_page},
                    )


# ============ HAVARD Börgen (clean table) ============
def parse_havard_borgen() -> Iterator[dict]:
    path = BASE / 'Havard' / 'Lista de precios Börgen - Marzo 2026 - USD.pdf'
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            pdf_page = pi + 1
            for tbl in page.extract_tables():
                if not tbl:
                    continue
                # Find header
                header_idx = None
                for hi, r in enumerate(tbl):
                    if r and any('SKU' in str(c or '').upper() for c in r):
                        header_idx = hi
                        break
                if header_idx is None:
                    continue
                for r in tbl[header_idx + 1:]:
                    if not r or len(r) < 5 or not r[0]:
                        continue
                    sku = (r[0] or '').strip().replace('\n', '')
                    desc = (r[1] or '').strip() if len(r) > 1 else ''
                    spec = (r[2] or '').strip() if len(r) > 2 else ''
                    iva = (r[3] or '').strip() if len(r) > 3 else ''
                    precio_raw = (r[4] or '').strip() if len(r) > 4 else ''
                    if not desc:
                        continue
                    m = re.search(r'([\d.,]+)', precio_raw)
                    price_num = to_money(m.group(1)) if m else None
                    currency = 'USD' if 'U$' in precio_raw.upper() or 'USD' in precio_raw.upper() else 'ARS'
                    parts = desc.split('\n')
                    producto = parts[0].strip()
                    rest = '\n'.join(p.strip() for p in parts[1:]) if len(parts) > 1 else ''
                    descripcion = (spec + ('\n' + rest if rest else '')).strip() or None
                    yield row(
                        'Havard',
                        Categoria='Börgen',
                        Codigo=sku,
                        Producto=producto,
                        Descripcion=descripcion,
                        Precio=price_num,
                        Moneda=currency,
                        IVA=iva,
                        **{'Archivo origen': path.name, 'Pagina PDF': pdf_page},
                    )


# ============ DANDA PDF (multi-line cells) ============
def parse_danda() -> Iterator[dict]:
    path = BASE / 'Danda' / 'Lista de Precios Marzo  2026.pdf'
    current_cat = None
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages, start=1):
            for tbl in page.extract_tables():
                if not tbl:
                    continue
                for r in tbl:
                    if not r:
                        continue
                    # Row can be a category header or data
                    col_a = (r[0] or '').strip()
                    col_b = (r[1] or '').strip() if len(r) > 1 else ''
                    if col_b == '' and col_a:
                        # Category row (e.g., "Cocinas\n(Exterior acero inoxidable...)")
                        cat_parts = [p.strip() for p in col_a.split('\n') if p.strip()]
                        if cat_parts:
                            current_cat = cat_parts[0]
                        continue
                    # Data row: col_a has newline-separated codes, col_b has newline-separated desc+price
                    codes = [c.strip() for c in col_a.split('\n') if c.strip()]
                    body_lines = [ln for ln in col_b.split('\n') if ln.strip()]
                    # Match lines to codes: prefer lines that contain a price. Lines with no $ are subcategory notes we skip.
                    data_lines = []
                    section = current_cat
                    for ln in body_lines:
                        if '$' in ln:
                            data_lines.append((section, ln))
                        else:
                            # Update section marker for following lines
                            # e.g. "LINEA SIENA", "LINEA MODENA (...)"
                            section = ln.strip()
                    # Match codes to data_lines by order (best-effort)
                    for i, (sec, ln) in enumerate(data_lines):
                        m = re.match(r'(.*?)\s*\$\s*([\d.,]+)\s*$', ln)
                        if not m:
                            continue
                        desc = m.group(1).strip()
                        price = to_money(m.group(2))
                        codigo = codes[i] if i < len(codes) else None
                        yield row(
                            'Danda',
                            Categoria=f'{current_cat} / {sec}' if sec and sec != current_cat else current_cat,
                            Codigo=codigo,
                            Producto=desc,
                            Precio=price,
                            Moneda='ARS',
                            IVA='Sin IVA',
                            **{'Archivo origen': path.name, 'Pagina PDF': pi},
                        )


# ============ Generic catalog-style HAVARD PDFs (text-block parsing) ============
def parse_havard_catalog_text(txt_path: Path, source_name: str, categoria: str) -> Iterator[dict]:
    """Parse Havard/Gelopar/FAME-style catalog text where each product has
    name + description lines + a row like 'SKU ...' followed by a price row
    'CODE  PRICE1  PRICE2...'."""
    paired = lines_with_pages(txt_path)
    lines = [ln for _, ln in paired]
    pages = [p for p, _ in paired]
    i = 0
    current = {}
    while i < len(lines):
        ln = lines[i]
        current_page = pages[i] if i < len(pages) else None
        # Look for SKU header line: contains 'SKU' and 'MAYORISTA' or 'CON IVA'
        if re.search(r'\bSKU\b', ln) and ('MAYORISTA' in ln.upper() or 'IVA' in ln.upper()):
            # The previous block (above) is the product name + description
            # Look at lines backwards to find product name
            desc_lines = []
            j = i - 1
            while j >= 0 and j > i - 20:
                prev = lines[j].strip()
                if not prev:
                    if desc_lines:
                        break
                    j -= 1
                    continue
                if re.search(r'\bSKU\b', prev) or re.match(r'^\s*STOCK\b', prev):
                    break
                desc_lines.insert(0, prev)
                j -= 1
            # Next few lines may contain the data (code + prices).
            # Skymsen SKUs always start with 987- (e.g. "987-EX", "987-DB-10T").
            # Tramontina rows in the same PDF start with an altura like "20mm" or "65mm"
            # followed by the Tramontina SKU — we must NOT match those.
            k = i + 1
            data_rows = []
            while k < len(lines) and k < i + 6:
                dl = lines[k].strip()
                if dl and re.match(r'^(?:STOCK\s+DISPONIBLE|SIN\s+STOCK)?\s*987-', dl, flags=re.I):
                    data_rows.append(dl)
                k += 1
            # Name = first non-empty desc line; rest = description
            name = ''
            description_parts = []
            for dl in desc_lines:
                if not name and not re.match(r'^[A-ZÁÉÍÓÚÜÑ0-9\s\-/\.]+$', dl):
                    # heuristic: first descriptive line
                    pass
                if not name:
                    name = dl
                else:
                    description_parts.append(dl)
            description = '\n'.join(description_parts) or None
            # parse data rows
            for dl in data_rows:
                # e.g. "987-EX             181.320                  266.999"
                dl_clean = re.sub(r'^(STOCK\s+DISPONIBLE|SIN\s+STOCK)\s*', '', dl, flags=re.I)
                m = re.match(r'^([\w\-]+)\s+([\d.,]+)(?:\s+([\d.,]+))?(?:\s+([\d.,]+))?', dl_clean)
                if not m:
                    continue
                sku = m.group(1)
                may = to_money(m.group(2))
                sug105 = to_money(m.group(3)) if m.group(3) else None
                sug21 = to_money(m.group(4)) if m.group(4) else None
                yield row(
                    'Havard',
                    Categoria=categoria,
                    Codigo=sku,
                    Producto=name or 'SIN NOMBRE',
                    Descripcion=description,
                    Precio=may,
                    Moneda='ARS',
                    IVA='Mayorista sin IVA',
                    Observaciones=' | '.join(x for x in [
                        f'sugerido +IVA 10.5%: {sug105:,.0f}' if sug105 else None,
                        f'sugerido +IVA 21%: {sug21:,.0f}' if sug21 else None,
                    ] if x) or None,
                    **{'Archivo origen': source_name, 'Pagina PDF': current_page},
                )
            i = k
            continue
        i += 1


# ============ RESINET (text) ============
def parse_resinet() -> Iterator[dict]:
    txt = (TXT / 'Lista precio ENERO 2026 RESINET.txt').read_text(encoding='utf-8', errors='replace')
    # Pattern: model lines like "R1", "R3", followed by description, and a price line "$96000"
    # pdfplumber gives us tables but text-based regex on the layout-extracted text is more robust here.
    # Strategy: find lines matching ^\s*R\d+\s*$ or ^\s*[A-Z]\d+$; group description lines; find next '$<num>'
    lines = [l for l in txt.splitlines()]
    blocks = []
    i = 0
    while i < len(lines):
        ln = lines[i].rstrip()
        m = re.match(r'^\s*([A-Z]{1,3}\d+)\s*$', ln)
        if m:
            code = m.group(1)
            desc = []
            price = None
            j = i + 1
            while j < len(lines) and j < i + 12:
                dl = lines[j].strip()
                pm = re.search(r'\$\s*([\d.,]+)', dl)
                if pm:
                    price = to_money(pm.group(1))
                    # text before price is description continuation
                    pre = dl[:pm.start()].strip()
                    if pre:
                        desc.append(pre)
                    i = j
                    break
                if dl and not re.match(r'^\s*[A-Z]{1,3}\d+\s*$', dl):
                    desc.append(dl)
                j += 1
            if price is not None and desc:
                yield row(
                    'Resinet',
                    Codigo=code,
                    Producto=desc[0],
                    Descripcion='\n'.join(desc[1:]) if len(desc) > 1 else None,
                    Precio=price,
                    Moneda='ARS',
                    IVA='Sin IVA',
                    **{'Archivo origen': 'Lista precio ENERO 2026 RESINET.pdf'},
                )
        i += 1


# ============ FADECO (text) ============
def parse_fadeco() -> Iterator[dict]:
    """FADECO PDF text has 3 product sections, each preceded by a centered title:
        'Cortadora de Fiambre' (con Modelo 330 / Modelo 300 sub-sections)
        'Picadora 32'
        'Amasadora'
    The section title is on its own line and applies until the next section.
    """
    txt_path = TXT / 'FADECO enero26.txt'
    lines = [l.rstrip() for l in txt_path.read_text(encoding='utf-8', errors='replace').splitlines()]
    SECTION_HEADERS = {
        'Cortadora de Fiambre': 'Cortadoras de Fiambre',
        'Picadora 32': 'Picadoras',
        'Amasadora': 'Amasadoras',
    }
    categoria = None
    i = 0
    while i < len(lines):
        ln = lines[i].strip()
        # Detect section header: exact stripped match
        if ln in SECTION_HEADERS:
            categoria = SECTION_HEADERS[ln]
            i += 1
            continue
        # Product line: "Fadeco <model> ...   $price"
        m = re.match(r'^(Fadeco\s+\S+.*?)\s+\$\s*([\d.,]+)\s*$', ln)
        if m:
            producto = m.group(1).strip()
            price = to_money(m.group(2))
            desc = None
            if i + 1 < len(lines) and lines[i + 1].strip().startswith('('):
                desc = lines[i + 1].strip()
            yield row(
                'Fadeco',
                Categoria=categoria,
                Producto=producto,
                Descripcion=desc,
                Precio=price,
                Moneda='ARS',
                IVA='Sin IVA',
                Observaciones='IVA 10.5% - Concesionario',
                **{'Archivo origen': 'FADECO enero26.pdf', 'Pagina PDF': 1},
            )
            i += 1
            continue
        # Other product lines: "Picadora 32, 2 HP ... $1.380.000" / "Amasadora 20 kilos  $1.010.000"
        m2 = re.match(r'^(Picadora\s+\d+.*?|Amasadora\s+.*?)\s+\$\s*([\d.,]+)\s*$', ln)
        if m2:
            producto = m2.group(1).strip()
            # Pick up RPM/HP spec from the next 1-2 lines if it looks like a spec
            desc_parts = []
            j = i + 1
            while j < len(lines) and j < i + 3:
                nxt = lines[j].strip()
                if not nxt or nxt.startswith('Picadora') or nxt.startswith('Amasadora') or nxt.startswith('Fadeco') or nxt in SECTION_HEADERS or '$' in nxt:
                    break
                if any(k in nxt.lower() for k in ('rpm', 'hp', 'parada', 'emergencia', 'mano', 'kilo')):
                    desc_parts.append(nxt)
                j += 1
            yield row(
                'Fadeco',
                Categoria=categoria,
                Producto=producto,
                Descripcion=' | '.join(desc_parts) or None,
                Precio=to_money(m2.group(2)),
                Moneda='ARS',
                IVA='Sin IVA',
                **{'Archivo origen': 'FADECO enero26.pdf', 'Pagina PDF': 1},
            )
        i += 1


# ============ DON SEGURA Amasadoras (text) ============
def parse_donsegura_amasadoras() -> Iterator[dict]:
    txt_path = TXT / 'Don Segura__AMASADORAS Y SOBADORAS MAYORISTA 01-05-26.txt'
    categoria = None
    for ln in txt_path.read_text(encoding='utf-8', errors='replace').splitlines():
        line = ln.strip()
        if not line:
            continue
        if re.match(r'^[A-ZÑÁÉÍÓÚ\s\-]+$', line) and len(line) < 60 and '$' not in line and 'LISTA' not in line and 'MAYORISTA' != line.strip() and not line.startswith('REGULACION') and not line.startswith('PRECIOS') and not line.startswith('DESCUENTOS'):
            # Likely a category header
            if line in ('AMASADORAS', 'SOBADORAS PASTELERAS') or 'SOBADORA' in line or 'LAMINADORA' in line:
                categoria = line
                continue
        m = re.match(r'^(.*?)\s+\$\s*([\d.,]+)\s*$', line)
        if m:
            producto = m.group(1).strip()
            price = to_money(m.group(2))
            if producto and price:
                yield row(
                    'Don Segura',
                    Categoria=categoria or 'Amasadoras/Sobadoras',
                    Producto=producto,
                    Precio=price,
                    Moneda='ARS',
                    IVA='Sin IVA (+10.5%)',
                    **{'Archivo origen': 'AMASADORAS Y SOBADORAS MAYORISTA 01-05-26.pdf', 'Pagina PDF': 1},
                )


# ============ DON SEGURA Mesas y Piletas (pdfplumber posicional) ============
def parse_donsegura_mesas() -> Iterator[dict]:
    """Parse 'LISTA DE PRECIOS MESAS Y PILETAS' using positional words.

    The PDF has 3 columns of mesas (SIN estante / C/estante / C/estante+zócalo).
    Column boundaries (from inspection):
        col 1: x0 in [85, 225]
        col 2: x0 in [230, 370]
        col 3: x0 in [375, 520]
    Each row has the dimension + 'U$S NN' in each of the 3 columns.
    PILETAS section is below, single column with 'XX x XX x XX cm fondo  U$S NN'.
    """
    src = BASE / 'Don Segura' / 'LISTA DE PRECIOS MESAS Y PILETAS - 7-04-26.pdf'
    if not src.exists():
        return
    archivo = 'LISTA DE PRECIOS MESAS Y PILETAS - 7-04-26.pdf'
    CATS = [
        'Mesa Inoxidable SIN estante 1,2mm (430)',
        'Mesa Inoxidable C/estante 1,2mm (430)',
        'Mesa C/estante y zócalo 10cm',
    ]
    col_bounds = [(85, 225), (230, 370), (375, 520)]
    with pdfplumber.open(src) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            # Group words by y-bucket (row)
            rows_by_top: dict[int, list] = {}
            for w in words:
                bucket = round(w['top'] / 3) * 3
                rows_by_top.setdefault(bucket, []).append(w)
            for top in sorted(rows_by_top):
                row_words = sorted(rows_by_top[top], key=lambda w: w['x0'])
                # Split into the 3 columns
                cols = [[], [], []]
                for w in row_words:
                    for ci, (lo, hi) in enumerate(col_bounds):
                        if lo <= w['x0'] < hi:
                            cols[ci].append(w)
                            break
                # For each column, expect: <DIM> x <DIM> m  U$S <price>
                for ci, col_words in enumerate(cols):
                    if not col_words:
                        continue
                    txt = ' '.join(w['text'] for w in col_words)
                    # Mesa: "1 x 60 m  U$S 88" or "1,20 x 60 m  U$S 115"
                    m = re.match(r'^([\d,]+)\s+x\s+(\d+)\s+m\s+U\$S\s+(\d+)\s*$', txt)
                    if m:
                        dim = f'{m.group(1)} x {m.group(2)} m'
                        price = to_money(m.group(3))
                        if price and price > 30:
                            yield row(
                                'Don Segura',
                                Categoria=CATS[ci],
                                Producto=f'Mesa {dim}',
                                Precio=price,
                                Moneda='USD',
                                IVA=None,
                                **{'Archivo origen': archivo},
                            )
                        continue
                    # Pileta: "50 x 30 x 30 cm fondo U$S 32"
                    m2 = re.search(r'(\d+)\s*[Xx]\s*(\d+)\s*[Xx]\s*(\d+)\s*cm\s*fondo\s*U\$S\s*(\d+)', txt, flags=re.I)
                    if m2:
                        dim = f'{m2.group(1)} x {m2.group(2)} x {m2.group(3)} cm'
                        price = to_money(m2.group(4))
                        if price and price > 10:
                            yield row(
                                'Don Segura',
                                Categoria='Piletas para soldar a mesa',
                                Producto=f'Pileta {dim} fondo',
                                Precio=price,
                                Moneda='USD',
                                IVA=None,
                                **{'Archivo origen': archivo},
                            )


# ============ DS Campanas (simple list from text) ============
def parse_ds_campanas() -> Iterator[dict]:
    txt_path = TXT / 'DS Campanas__LISTA DE PRECIOS DS CAMPANAS.txt'
    text = txt_path.read_text(encoding='utf-8', errors='replace')
    lines = text.splitlines()
    results = []
    section = '3 velocidades'
    for ln in lines:
        l = ln.strip()
        # Section header detection
        if re.match(r'^Campanas?\s+\d+\s+velocidad', l, re.IGNORECASE):
            m_sec = re.match(r'^(Campanas?\s+\d+\s+velocidad(?:es)?)', l, re.IGNORECASE)
            if m_sec:
                section = m_sec.group(1).lower().replace('campanas', 'campana')
            continue
        m = re.search(r'(Campanas?\s+\d+\s*cm\.?)\s*\$\s*([\d.,]+)', ln)
        if m:
            results.append(row(
                'DS Campanas',
                Categoria=f'Slim / Semi circulares / Piramidales - {section}',
                Producto=m.group(1).strip(),
                Precio=to_money(m.group(2)),
                Moneda='ARS',
                IVA=None,
                **{'Archivo origen': 'LISTA DE PRECIOS DS CAMPANAS.pdf', 'Pagina PDF': 1},
            ))
    return iter(results)


# ============ TURBOBLENDER CATALOGO DF V.02.25 (text, no prices in most) ============
def parse_turboblender_catalogo() -> Iterator[dict]:
    """This catalog has product names and codes but prices are in the separate (image-only) price list.
    Extract names + codes for now; prices left empty and flagged.
    """
    txt_path = TXT / 'Turboblender__CATALOGO DF V.02.25.txt'
    lines = [l.rstrip() for l in txt_path.read_text(encoding='utf-8', errors='replace').splitlines()]
    # This catalog probably has product codes like "DF" and names. Without seeing structure, we do minimal parse.
    # Flag whole catalog as "para revisar" — return empty.
    return iter([])


# ============ MB CATALOGO ============
def parse_mb_catalogo() -> Iterator[dict]:
    """Parse 'CATALOGO DE PRODUCTOS MB.pdf' positional via pdfplumber.

    Layout: each page has 1-3 product photos with a name above and '$ NNN.NNN' below.
    Strategy: extract words, group lines by 'top' coord (5px tolerance), find lines
    containing a '$' token, pair with the closest non-price text block above.
    """
    src = BASE / 'CATALOGO DE PRODUCTOS MB.pdf'
    if not src.exists():
        return
    seen = set()
    with pdfplumber.open(src) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            # Group by top coord
            rows_by_top: dict[float, list] = {}
            for w in words:
                bucket = None
                for k in rows_by_top:
                    if abs(w['top'] - k) <= 4:
                        bucket = k
                        break
                if bucket is None:
                    bucket = w['top']
                rows_by_top.setdefault(bucket, []).append(w)
            sorted_tops = sorted(rows_by_top)
            line_texts = []
            for t in sorted_tops:
                ws = sorted(rows_by_top[t], key=lambda w: w['x0'])
                line_texts.append((t, ' '.join(w['text'] for w in ws).strip()))
            for idx, (t, text) in enumerate(line_texts):
                m = re.search(r'\$\s*([\d.,]+)', text)
                if not m:
                    continue
                price = to_money(m.group(1))
                if not price or price < 1000:
                    continue
                # Name = previous non-price line that looks like a product (3+ chars, not all numeric)
                name = None
                for back_idx in range(idx - 1, max(idx - 6, -1), -1):
                    candidate = line_texts[back_idx][1]
                    if '$' in candidate or not candidate or len(candidate) < 4:
                        continue
                    # Avoid lines that are mostly garbage/encoded
                    if candidate.lower() in {'cm', 'mm', 'kg'}:
                        continue
                    # Skip lines that are duplicate words like "Quemador   Quemador"
                    tokens = candidate.split()
                    if len(tokens) >= 2 and tokens[0] == tokens[1]:
                        candidate = ' '.join(tokens[len(tokens)//2:])
                    name = candidate
                    break
                if not name:
                    continue
                # Same-line text before the $ may be the name
                before = text[:m.start()].strip()
                if before and len(before) > 3:
                    name = before
                # Dedupe by (name, price)
                key = (name, price)
                if key in seen:
                    continue
                seen.add(key)
                yield row(
                    'MB',
                    Categoria='MB',
                    Producto=name,
                    Precio=price,
                    Moneda='ARS',
                    IVA=None,
                    **{'Archivo origen': 'CATALOGO DE PRODUCTOS MB.pdf'},
                )


# ============ DUNIA CATALOGO PDF (text) ============
def parse_dunia_catalogo_pdf() -> Iterator[dict]:
    """Dunia PDF catalog — most items in the XLSX already; extract code+name+price from PDF text
    as a fallback for items not in XLSX."""
    txt_path = TXT / 'Dunia__CATALOGO_DUNIA.txt'
    if not txt_path.exists():
        return iter([])
    lines = [l.rstrip() for l in txt_path.read_text(encoding='utf-8', errors='replace').splitlines()]
    results = []
    for ln in lines:
        # Look for price patterns "$ XXX.XXX" with a product description
        m = re.search(r'\$\s*([\d.,]+)', ln)
        if m:
            price = to_money(m.group(1))
            before = ln[:m.start()].strip()
            if before and len(before) > 5 and price and price > 100:
                results.append(row(
                    'Dunia',
                    Categoria='(Desde PDF catálogo)',
                    Producto=before,
                    Precio=price,
                    Moneda='ARS',
                    **{'Archivo origen': 'CATÁLOGO DUNIA.pdf'},
                ))
    return iter(results)


# ============ DUNIA EQUIPAMIENTO ============
def parse_dunia_equipamiento() -> Iterator[dict]:
    txt_path = TXT / 'Dunia__EQUIPAMIENTO 2025.txt'
    lines = [l.rstrip() for l in txt_path.read_text(encoding='utf-8', errors='replace').splitlines()]
    results = []
    for ln in lines:
        m = re.search(r'\$\s*([\d.,]+)', ln)
        if m:
            price = to_money(m.group(1))
            before = ln[:m.start()].strip()
            if before and len(before) > 5 and price and price > 100:
                results.append(row(
                    'Dunia',
                    Categoria='Equipamiento 2025',
                    Producto=before,
                    Precio=price,
                    Moneda='ARS',
                    **{'Archivo origen': 'EQUIPAMIENTO 2025.pdf'},
                ))
    return iter(results)


# ============ HAVARD Mayorista-marzo-sinprecios (text, no prices) ============
# Skip — just catalog data without prices; flag as needing prices


def _ar_money(s):
    """Convert Argentine-formatted numeric string to float.

    Treats '.' as thousands separator (always) and ',' as decimal.
    Use for new OCR parsers where 'to_money' miscategorizes '5.194' as 5.194 instead of 5194.
    """
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    s = re.sub(r'(?i)usd|u\$s|u\$d|u\$|\$|ar\$', '', s).strip().replace(' ', '')
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace('.', '')
    try:
        return float(s)
    except ValueError:
        return None


# ============ TURBOBLENDER De Francesco (OCR de LISTA DE PRECIOS DF V.01.26) ============
def parse_turboblender_df_pricelist() -> Iterator[dict]:
    """Parse OCR output of 'LISTA DE PRECIOS DF V.01.26.pdf' (De Francesco / Turboblender).

    Each product page ends with a table whose rows have the shape:
        DF-CODE  <attrs...>  USD <distribuidor>  USD <precio sugerido>
    Categoría se infiere de la línea-título previa en mayúsculas (HELADERA EXHIBIDORA, MOSTRADOR, FREEZER, etc.).
    """
    txt_path = TXT_OCR / 'Turboblender__LISTA DE PRECIOS DF V.01.26.txt'
    if not txt_path.exists():
        return iter([])
    paired = lines_with_pages(txt_path)
    lines = [ln for _, ln in paired]
    pages = [p for p, _ in paired]
    cat_keys = (
        'HELADERA', 'MOSTRADOR', 'FREEZER', 'EXHIBIDORA', 'VITRINA', 'CAVA',
        'CONSERVADOR', 'CHOCOLATER', 'ABATIDOR', 'BOTELLER', 'AUTOSERVICIO',
        'CERVECER', 'BACHA', 'BARTENDER', 'GASTRONOM', 'PIZZER', 'TURRONER',
        'CARNICER', 'CARNICERIA', 'PESCADER', 'CALENTADOR', 'FRIO',
    )
    data_re = re.compile(
        r'(DF-[A-Z0-9]+(?:\s?[A-Z]{1,4})?)\s+(.+?)\s+USD\s*([\d.,]+)\s*[\|¡:]?\s*USD\s*([\d.,]+)',
        re.IGNORECASE,
    )
    seen_codes = set()
    current_cat = None
    results = []
    for idx, raw in enumerate(lines):
        ln = raw.strip()
        if not ln:
            continue
        cur_page = pages[idx] if idx < len(pages) else None
        # Detect category — uppercase title containing one of the keywords
        if any(k in ln.upper() for k in cat_keys) and ln.upper() == ln and len(ln) < 80 and 'DF-' not in ln and 'USD' not in ln:
            letters = sum(1 for c in ln if c.isalpha())
            if letters >= 6:
                current_cat = ln.strip(' :|')
                continue
        for m in data_re.finditer(raw):
            code = re.sub(r'\s+', '', m.group(1)).upper()
            mid = m.group(2).strip(' :|¡')
            dist = _ar_money(m.group(3))
            sugerido = _ar_money(m.group(4))
            if not dist or dist < 50:
                continue
            key = (code, dist)
            if key in seen_codes:
                continue
            seen_codes.add(key)
            obs = f'precio sugerido al público (sin IVA): USD {sugerido:,.2f}' if sugerido else None
            results.append(row(
                'Turboblender',
                Categoria=current_cat or 'De Francesco',
                Codigo=code,
                Producto=code,
                Descripcion=mid or None,
                Precio=dist,
                Moneda='USD',
                IVA='Sin IVA',
                Observaciones=obs,
                **{'Archivo origen': 'LISTA DE PRECIOS DF V.01.26.pdf', 'Pagina PDF': cur_page},
            ))
    return iter(results)


# ============ TURBOBLENDER Línea TB (OCR de 'lista abril 26') ============
def parse_turboblender_lista_abril() -> Iterator[dict]:
    """Parse OCR output of 'lista abril 26.pdf' (línea Turboblender — licuadoras, mixers, etc.).

    Cada producto tiene un encabezado tipo 'Licuadora Profesional TB-020 ...' y, varias líneas más
    abajo, una línea con 'Mayorista $XXX,XX  Público $YYY,YY'. El OCR captura solo algunas debido
    al estilo gráfico de los precios. Emitimos solo filas con al menos un precio numérico.
    """
    txt_path = TXT_OCR / 'Turboblender__Lista Turboblender abril 26.txt'
    if not txt_path.exists():
        return iter([])
    paired = lines_with_pages(txt_path)
    lines = [ln for _, ln in paired]
    pages = [p for p, _ in paired]
    prod_re = re.compile(
        r'((?:Licuadora|Mixer|Exprimidor|Procesadora|Cutter|Picadora|Cortadora|Cafetera|'
        r'Anafe|Freidora|Plancha|Waflera|Wafleras|Tostadora|Tostador|Horno|Calentador|'
        r'Fabricadora|Microondas|Accesorio|Enjuagador|Afilador)\b[^\n]*?\b(TB-[A-Z0-9]+(?:\s?[A-Z0-9]+)?))',
        re.IGNORECASE,
    )
    results = []
    current_name = None
    current_code = None
    current_page = None
    for idx, ln in enumerate(lines):
        cur_page = pages[idx] if idx < len(pages) else None
        m = prod_re.search(ln)
        if m:
            current_name = re.sub(r'\s+', ' ', m.group(1)).strip()
            current_code = re.sub(r'\s+', '', m.group(2)).upper()
            current_page = cur_page
            continue
        if not current_code:
            continue
        if 'MAYORISTA' not in ln.upper() and 'PÚBLICO' not in ln.upper() and 'PUBLICO' not in ln.upper():
            continue
        nums = re.findall(r'\$\s*([\d][\d.,]*)', ln)
        nums = [_ar_money(n) for n in nums]
        nums = [n for n in nums if n and n > 1000]
        if not nums:
            continue
        mayorista = nums[0]
        publico = nums[1] if len(nums) > 1 else None
        upper = ln.upper()
        if 'MAYORISTA' not in upper and ('PÚBLICO' in upper or 'PUBLICO' in upper):
            mayorista, publico = None, nums[0]
        obs_parts = []
        if publico:
            obs_parts.append(f'Público IVA incl: ${publico:,.2f}')
        if mayorista is None:
            obs_parts.append('mayorista no legible (OCR)')
        results.append(row(
            'Turboblender',
            Categoria='Línea Turboblender',
            Codigo=current_code,
            Producto=current_name,
            Precio=mayorista if mayorista is not None else publico,
            Moneda='ARS',
            IVA='Sin IVA' if mayorista is not None else 'IVA incluido',
            Observaciones=' | '.join(obs_parts) or None,
            **{'Archivo origen': 'Lista Turboblender abril 26.pdf', 'Pagina PDF': current_page},
        ))
        current_code = None
    return iter(results)


# ============ HAVARD Preventa Calefacción Börgen (OCR) ============
def parse_havard_preventa_borgen() -> Iterator[dict]:
    """Parse OCR output of 'PREVENTA CALEFACCION BORGEN 2026 -.pdf'."""
    txt_path = TXT_OCR / 'Havard__PREVENTA CALEFACCION BORGEN 2026.txt'
    if not txt_path.exists():
        return iter([])
    txt = txt_path.read_text(encoding='utf-8', errors='replace')
    # Productos visibles (3 paneles):
    # 1) PANEL CONVECTOR DIGITAL CON FORZADOR     -> $110.576 + IVA
    # 2) PANEL CONVECTOR ANALÓGICO CON FORZADOR   -> precio ilegible (OCR roto)
    # 3) PANEL CONVECTOR 1200 WATTS BORGEN CON-71200 -> $93.872 + IVA
    results = []
    paired = lines_with_pages(txt_path)
    blocks = re.split(r'(?=^PANEL\s+CONVECTOR)', txt, flags=re.MULTILINE)
    for blk in blocks:
        title_m = re.search(r'(PANEL\s+CONVECTOR[^\n]*)', blk)
        if not title_m:
            continue
        title = title_m.group(1).strip()
        code_m = re.search(r'BORGEN\s+([A-Z0-9\-]+)', blk)
        code = code_m.group(1).rstrip('/') if code_m else None
        price_m = re.search(r'\$\s*([\d]{1,3}(?:[.,]\d{3})+)', blk)
        if not price_m:
            continue
        price = _ar_money(price_m.group(1))
        if not price or price < 1000:
            continue
        # Find page where this product's title appears
        product_page = None
        for pp, ln in paired:
            if title[:30] in ln:
                product_page = pp
                break
        results.append(row(
            'Havard',
            Categoria='Calefacción Börgen (preventa)',
            Codigo=code,
            Producto=title,
            Precio=price,
            Moneda='ARS',
            IVA='Sin IVA',
            Observaciones='Preventa con 4 e-checks fijos. Envío gratis. Mín 30 unidades.',
            **{'Archivo origen': 'PREVENTA CALEFACCION BORGEN 2026 -.pdf', 'Pagina PDF': product_page},
        ))
    return iter(results)


# ============ BIANCHI Cortadora NOVA 330 (OCR) ============
def parse_bianchi_nova330() -> Iterator[dict]:
    """Parse OCR output of 'Cortadora BIANCHI NOVA  330.pdf'."""
    txt_path = TXT_OCR / 'Bianchi__Cortadora NOVA 330.txt'
    if not txt_path.exists():
        return iter([])
    txt = txt_path.read_text(encoding='utf-8', errors='replace')
    m = re.search(r'\$\s*([\d]{1,3}(?:[.,]\d{3})+)', txt)
    if not m:
        return iter([])
    price = _ar_money(m.group(1))
    if not price or price < 10000:
        return iter([])
    return iter([row(
        'Bianchi',
        Categoria='Cortadoras de Fiambre',
        Codigo='NOVA 330',
        Producto='Cortadora de Fiambre BIANCHI NOVA 330',
        Descripcion=('Construcción en aluminio pulido anodizado. Cuchilla de acero inoxidable Ø 330 mm. '
                     'Motor monofásico 1/3 HP - 220V. Transmisión a correa. Patas antideslizantes. '
                     'Regulador de espesor de corte. Dimensiones 42x42x57 cm.'),
        Precio=price,
        Moneda='ARS',
        IVA=None,
        **{'Archivo origen': 'Cortadora BIANCHI NOVA  330.pdf', 'Pagina PDF': 1},
    )])


# ============ HAVARD FAME (PDF positional) ============
def parse_havard_fame_pdf() -> Iterator[dict]:
    """Extract products from 'FAME HAVARD MARZO 2026.pdf' using pdfplumber positional words.

    Layout (per row):
        x0 < 140  -> Código
        140..395  -> Descripción (con asterisco '*' que marca IVA reducido)
        x0 >= 395 -> Precio (formato AR: '1.896.188,24')
    Filas sin precio o sin código se ignoran. Categorías se infieren de filas
    "solo descripción" en mayúsculas.
    """
    src = BASE / 'Havard' / 'FAME HAVARD MARZO 2026.pdf'
    if not src.exists():
        return
    archivo = 'FAME HAVARD MARZO 2026.pdf'
    current_cat = 'FAME'
    with pdfplumber.open(src) as pdf:
        for pi, page in enumerate(pdf.pages, start=1):
            words = page.extract_words()
            # cluster by row top (3px tolerance)
            rows: dict[float, list] = {}
            for w in words:
                bucket = None
                for k in rows:
                    if abs(w['top'] - k) <= 3:
                        bucket = k
                        break
                if bucket is None:
                    bucket = w['top']
                rows.setdefault(bucket, []).append(w)
            for top in sorted(rows):
                row_words = sorted(rows[top], key=lambda w: w['x0'])
                code_w = [w for w in row_words if w['x0'] < 140]
                star = any(w['text'] == '*' for w in row_words if w['x0'] < 200)
                desc_w = [w for w in row_words if 140 <= w['x0'] < 395 and w['text'] != '*']
                price_w = [w for w in row_words if w['x0'] >= 395]
                code = ' '.join(w['text'] for w in code_w).strip()
                desc = ' '.join(w['text'] for w in desc_w).strip()
                prc = ' '.join(w['text'] for w in price_w).strip()
                # Skip headers/footers
                if not (code or desc or prc):
                    continue
                if 'No incluyen IVA' in desc or 'No incluyen IVA' in prc:
                    continue
                if desc.lower().startswith('código') or desc.lower().startswith('descripción'):
                    continue
                # category-only row
                if desc and not prc and not code:
                    if 5 < len(desc) < 80:
                        current_cat = f'FAME · {desc}'
                    continue
                if not prc or not re.search(r'\d', prc):
                    continue
                price = _ar_money(prc)
                if price is None or price < 100:
                    continue
                if not code:
                    code = ''
                # Split desc into producto/descripcion: first ~40 chars as producto
                producto = desc if len(desc) <= 60 else desc[:60].rsplit(' ', 1)[0]
                descripcion = '' if producto == desc else desc
                iva = 21.0 if star else 10.5
                yield row(
                    'Havard',
                    Categoria=current_cat,
                    Codigo=code,
                    Producto=producto or code,
                    Descripcion=descripcion,
                    Precio=price,
                    Moneda='ARS',
                    IVA=iva,
                    Observaciones='IVA 21% (accesorio/repuesto)' if star else 'IVA 10,5% (máquina)',
                    **{'Archivo origen': archivo, 'Pagina PDF': pi},
                )


# ============ HAVARD GELOPAR (PDF text per-page) ============
def parse_havard_gelopar_pdf() -> Iterator[dict]:
    """Extract products from 'GeloparAbril2026.pdf' (USD).

    Each page contains a category heading, a 'GELOPAR <model>' line and a SKU+price
    line of the form '<sku>  USD <price>'. SKU pattern: 374-XXXXX.
    """
    src = BASE / 'Havard' / 'GeloparAbril2026.pdf'
    if not src.exists():
        return
    archivo = 'GeloparAbril2026.pdf'
    cat_re = re.compile(
        r'^(FREEZER|HELADERA|EXHIBIDORA|CONSERVADORA|VITRINA|MOSTRADOR|CHOCOLATER|CARNICER|'
        r'CAVA|SHOWCASE|CONGELAD|EXPOSITOR|HORIZONT|VERTICAL|BOTELLER|REFRIGERAD|CONGELAD)',
        re.IGNORECASE,
    )
    sku_re = re.compile(r'(\d{3,4}-[A-Z0-9\-]+)\s+USD\s*([\d.,]+)')
    with pdfplumber.open(src) as pdf:
        for pi, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ''
            last_type = None
            last_gelopar = None
            for raw in text.splitlines():
                ln = raw.strip()
                if not ln:
                    continue
                ln_up = ln.upper()
                if 'GELOPAR' in ln_up and ln_up.startswith('GELOPAR '):
                    last_gelopar = ln
                    continue
                if cat_re.match(ln_up) and 'GELOPAR' not in ln_up and 'USD' not in ln_up:
                    last_type = ln
                m = sku_re.search(ln)
                if m:
                    sku = m.group(1).strip()
                    p_raw = m.group(2)
                    if ',' in p_raw and '.' in p_raw:
                        # AR format e.g. 1.963,92
                        price = float(p_raw.replace('.', '').replace(',', '.'))
                    elif ',' in p_raw:
                        price = float(p_raw.replace(',', '.'))
                    else:
                        price = float(p_raw)
                    producto = last_gelopar or (last_type or 'Gelopar')
                    yield row(
                        'Havard',
                        Categoria=f'Gelopar · {last_type}' if last_type else 'Gelopar',
                        Codigo=sku,
                        Producto=producto,
                        Descripcion=last_type or '',
                        Precio=price,
                        Moneda='USD',
                        IVA=10.5,
                        **{'Archivo origen': archivo, 'Pagina PDF': pi},
                    )


# ============ RESINET (PDF positional) ============
def parse_resinet_pdf() -> Iterator[dict]:
    """Extract products from 'Lista precio ENERO 2026 RESINET.pdf' using positional words.

    Layout per page (from page 2 onward):
        Modelo (x<200)  |  Código  |  Descripción (240..480)  |  Foto  |  Precio (x>=700, '$NNNNN')
    """
    src = BASE / 'Lista precio ENERO 2026 RESINET.pdf'
    if not src.exists():
        return
    archivo = 'Lista precio ENERO 2026 RESINET.pdf'
    skip_codes = {'Modelo', 'Código', 'Descripción', 'Foto', 'Precio'}
    with pdfplumber.open(src) as pdf:
        for pi in range(1, len(pdf.pages)):
            words = pdf.pages[pi].extract_words()
            prices = [w for w in words if w['text'].startswith('$')]
            codes_raw = [
                w for w in words
                if w['x0'] < 200 and w['top'] > 30 and w['text'] not in skip_codes
            ]
            codes_raw.sort(key=lambda w: (w['top'], w['x0']))
            # Continuation words: when these appear in the code column on a
            # subsequent line, merge them with the previous code (e.g. ESQUEL
            # then CON BASE = ESQUEL CON BASE).
            CONTINUATIONS = {'BASE', 'CON', 'TAPA', 'T', 'Y', 'y', 'GRILL', 'PRO'}
            merged: list[dict] = []
            for w in codes_raw:
                if not merged:
                    merged.append(dict(w))
                    continue
                last = merged[-1]
                # Horizontal continuation (same line)
                if abs(w['top'] - last['top']) < 5 and (w['x0'] - last['x1']) < 30:
                    last['text'] += ' ' + w['text']
                    last['x1'] = w['x1']
                    continue
                # Vertical continuation: short distance below (<= 18px), x0 near same column,
                # and the word looks like a SKU continuation (uppercase, short, common token)
                if (
                    0 < w['top'] - last['top'] <= 18
                    and abs(w['x0'] - last['x0']) < 40
                    and (w['text'] in CONTINUATIONS or (w['text'].isupper() and len(w['text']) <= 5))
                ):
                    last['text'] += ' ' + w['text']
                    last['top'] = w['top']  # so the descriptor block follows from the merged top
                    last['x1'] = max(last['x1'], w['x1'])
                    continue
                merged.append(dict(w))
            descs = [
                w for w in words
                if 240 < w['x0'] < 480 and w['top'] > 30 and w['text'] != 'Descripción'
            ]
            for pr in prices:
                candidates = [c for c in merged if c['top'] <= pr['top'] + 5]
                if not candidates:
                    continue
                code = max(candidates, key=lambda c: c['top'])
                later = [c for c in merged if c['top'] > code['top'] + 5]
                top_max = min((c['top'] for c in later), default=9999)
                block = sorted(
                    [d for d in descs if code['top'] - 3 <= d['top'] < top_max],
                    key=lambda d: (d['top'], d['x0']),
                )
                desc_text = ' '.join(d['text'] for d in block).strip()
                price_str = pr['text'].replace('$', '').replace('.', '').replace(',', '')
                try:
                    price = float(price_str)
                except ValueError:
                    continue
                if price < 100:
                    continue
                code_txt = code['text'].strip()
                # Repuestos (page index >= 14 typically) -> IVA 21%
                iva = 21.0 if code_txt.upper().startswith('REPUESTO') or pi >= 14 else 10.5
                yield row(
                    'Resinet',
                    Categoria='Repuestos' if iva == 21.0 else 'Máquinas',
                    Codigo=code_txt,
                    Producto=desc_text[:80] if desc_text else code_txt,
                    Descripcion=desc_text,
                    Precio=price,
                    Moneda='ARS',
                    IVA=iva,
                    **{'Archivo origen': archivo, 'Pagina PDF': pi + 1},
                )


# ============ ARE (PRECIOS) ============
def parse_are_precios() -> Iterator[dict]:
    """Parse 'ARE/PRECIOS.pdf' - text-extractable, format 'PRODUCT ... $ 103,884' (comma = thousands)."""
    txt_path = TXT / 'ARE__PRECIOS.txt'
    if not txt_path.exists():
        return
    text = txt_path.read_text(encoding='utf-8', errors='replace')
    # Names may wrap across lines, with the $ price appearing on a later line.
    # Strategy: scan lines for a name-like prefix; capture the next $price within the next 2 lines.
    lines = [l for l in text.splitlines()]
    i = 0
    # Determine category from product type word
    def cat_for(name: str) -> str:
        u = name.upper()
        if u.startswith('ANAFE DOBLE'):
            return 'Anafes dobles'
        if u.startswith('ANAFE'):
            return 'Anafes'
        if u.startswith('MECHERO'):
            return 'Mecheros'
        if 'RALLADORA' in u:
            return 'Ralladoras'
        if 'APLANACARNES' in u:
            return 'Aplanacarnes'
        if 'SALAMANDRA' in u:
            return 'Salamandras'
        return 'ARE'

    while i < len(lines):
        ln = lines[i]
        # Name line: starts with an uppercase word (ANAFE/MECHERO/RALLADORA/etc) and has 3+ uppercase letters
        m_name = re.match(r'^\s*((?:ANAFE|MECHERO|RALLADORA|APLANACARNES|KIT|MECHERO).+?)(?:\s*\$\s*([\d.,]+))?\s*$', ln)
        if not m_name:
            # Inline at end-of-list (e.g. '50X50$129308-')
            m_inline = re.match(r'^([\dxX]+)\s*\$\s*(\d[\d.,]+)-?\s*$', ln.strip())
            if m_inline:
                name = f'Plancha {m_inline.group(1)}'
                price = to_money(m_inline.group(2))
                if price:
                    yield row(
                        'ARE',
                        Categoria='Planchas',
                        Producto=name,
                        Precio=price,
                        Moneda='ARS',
                        IVA=None,
                        **{'Archivo origen': 'PRECIOS.pdf', 'Pagina PDF': 1},
                    )
            i += 1
            continue
        name = re.sub(r'[.…·…\s]{2,}.*$', '', m_name.group(1)).strip()
        name = re.sub(r'\s+', ' ', name)
        price_str = m_name.group(2)
        # If price not on same line, look at next 2 lines for "$ NNN,NNN"
        if not price_str:
            for j in range(i + 1, min(i + 3, len(lines))):
                m_price = re.search(r'\$\s*\.?([\d.,]+)', lines[j])
                if m_price:
                    price_str = m_price.group(1)
                    i = j
                    break
        if price_str:
            price = to_money(price_str)
            if price and price > 1000:
                yield row(
                    'ARE',
                    Categoria=cat_for(name),
                    Producto=name,
                    Precio=price,
                    Moneda='ARS',
                    IVA=None,
                    Observaciones='IVA + 10,5%',
                    **{'Archivo origen': 'PRECIOS.pdf', 'Pagina PDF': 1},
                )
        i += 1


# ============ Picadoras y sierras (Asadores DEC) ============
def parse_picadoras_sierras() -> Iterator[dict]:
    """Parse 'Picadoras y sierras.pdf'. Layout is irregular (price sometimes before,
    sometimes after the product name), so we collect all product names and all $prices
    in source order and zip them.
    """
    txt_path = TXT / 'Picadoras y sierras.txt'
    if not txt_path.exists():
        return
    text = txt_path.read_text(encoding='utf-8', errors='replace')
    name_re = re.compile(r'^\s*((?:Picadora|Sierra|Cortadora)[^$\n]+)$', re.IGNORECASE)
    price_re = re.compile(r'\$\s*([\d.,]+)')
    names = []
    prices = []
    # Buffer of description lines per name
    descriptions: list[list[str]] = []
    pending_desc: list[str] = []
    for ln in text.splitlines():
        l = ln.strip()
        if not l:
            continue
        m_name = name_re.match(l)
        if m_name and 'industrial' not in l.lower() or (m_name and re.search(r'\bN[º°]\s*\d', l)):
            # Save pending description to previous name slot if any
            if names and len(descriptions) < len(names):
                descriptions.append(pending_desc[:])
                pending_desc = []
            names.append(m_name.group(1).strip())
            continue
        m_price = price_re.search(l)
        if m_price:
            # close pending description for last name
            if names and len(descriptions) < len(names):
                descriptions.append(pending_desc[:])
                pending_desc = []
            prices.append(m_price.group(1))
            continue
        # Otherwise, description for the most recently seen product
        pending_desc.append(l)
    # Flush last desc
    if names and len(descriptions) < len(names):
        descriptions.append(pending_desc[:])
    # Re-extract names with simpler regex (the conditional in the loop was buggy)
    names = []
    for ln in text.splitlines():
        l = ln.strip()
        if re.match(r'^(Picadora|Sierra|Cortadora)\b', l, re.IGNORECASE) and '$' not in l:
            names.append(re.sub(r'\s+', ' ', l).strip())
    prices = [to_money(p) for p in price_re.findall(text)]
    prices = [p for p in prices if p and p > 100000]
    seen = set()
    for name, price in zip(names, prices):
        if (name, price) in seen:
            continue
        seen.add((name, price))
        # Best-effort page: search for the name in the per-page text. Falls back to 1.
        page_num = 1
        for pp, ln in lines_with_pages(txt_path):
            if name and name[:25].lower() in ln.lower():
                page_num = pp
                break
        yield row(
            'Asadores DEC',
            Categoria='Picadoras y sierras',
            Producto=name,
            Precio=price,
            Moneda='ARS',
            IVA=None,
            **{'Archivo origen': 'Picadoras y sierras.pdf', 'Pagina PDF': page_num},
        )


# ============ Calefactores de exterior (Asadores DEC) ============
def parse_calefactores_exterior() -> Iterator[dict]:
    txt_path = TXT / 'Calefactores de exterior.txt'
    if not txt_path.exists():
        return
    text = txt_path.read_text(encoding='utf-8', errors='replace')
    # Line-anchored: title is 'Calefactores de exterior <variant>' as its own line.
    # The header 'Calefactores de exterior' (no variant) on line 1 must not match.
    titles = re.findall(r'^\s*Calefactores de exterior\s+(\w+)\s*$', text, flags=re.MULTILINE)
    prices = [to_money(p) for p in re.findall(r'\$\s*([\d.,]+)', text)]
    prices = [p for p in prices if p and p > 10000]
    desc_common = ('Chapa de 20 pintada electrostática (negro, blanco o rojo). Gas envasado. '
                   'Incluye válvula de seguridad y piedra volcánica.')
    for variant, price in zip(titles, prices):
        yield row(
            'Asadores DEC',
            Categoria='Calefactores de exterior',
            Producto=f'Calefactor de exterior {variant}',
            Descripcion=desc_common,
            Precio=price,
            Moneda='ARS',
            IVA=None,
            Observaciones='No incluye garrafa ni regulador',
            **{'Archivo origen': 'Calefactores de exterior.pdf', 'Pagina PDF': 1},
        )


# ============ Mantenedores de calor (Asadores DEC) ============
def parse_mantenedores_calor() -> Iterator[dict]:
    txt_path = TXT / 'Mantenedores de calor.txt'
    if not txt_path.exists():
        return
    text = txt_path.read_text(encoding='utf-8', errors='replace')
    # Two products PRISMA 10 / PRISMA 6 with prices "$ 443.000" and "$408.000"
    # Find each PRISMA title and its associated price.
    # Simplest: find pairs of $prices, then back-locate product names.
    titles = re.findall(r'PRISMA\s+\d+\s+bandejas', text)
    prices = re.findall(r'\$\s*([\d.,]+)', text)
    desc_common = (
        'Cuerpo de acero inoxidable y vidrio templado con puertas corredizas. '
        'Bandejas de acero inoxidable. Tensión 220V, 300W. '
        'Termostato 0–90°C con corte automático.'
    )
    for title, price_raw in zip(titles, prices):
        price = to_money(price_raw)
        if not price or price < 10000:
            continue
        yield row(
            'Asadores DEC',
            Categoria='Mantenedores de calor',
            Producto=f'Mantenedor de calor {title.strip(":")}',
            Descripcion=desc_common,
            Precio=price,
            Moneda='ARS',
            IVA=None,
            **{'Archivo origen': 'Mantenedores de calor.pdf', 'Pagina PDF': 1},
        )


# ============ Bandejas LGR ============
def parse_bandejas_lgr() -> Iterator[dict]:
    txt_path = TXT / 'Bandejas LGR.txt'
    if not txt_path.exists():
        return
    paired = lines_with_pages(txt_path)
    current_cat = 'Bandejas'
    sub = None
    cat_keys = {'BANDEJAS': 'Bandejas', 'PIZZERAS': 'Pizzeras', 'HAMBURGUESERAS': 'Hamburgueseras', 'TRINCHA PAN': 'Trincha pan'}
    for page_n, ln in paired:
        l = ln.strip()
        if not l:
            continue
        for key, val in cat_keys.items():
            if l == key:
                current_cat = val
                sub = None
                break
        else:
            if l.startswith('Para carro'):
                sub = l
                continue
            m = re.match(r'^(.+?)\s+\$\s*([\d.,]+)\s+\$\s*([\d.,]+)\s*$', l)
            if not m:
                continue
            measure = m.group(1).strip()
            p_enlozada = to_money(m.group(2))
            p_chapa = to_money(m.group(3))
            cat = f'{current_cat} / {sub}' if sub else current_cat
            if p_enlozada and p_enlozada > 100:
                yield row(
                    'LGR',
                    Categoria=cat,
                    Producto=f'{current_cat[:-1] if current_cat.endswith("s") else current_cat} enlozada {measure}',
                    Descripcion='Enlozada' + (' - chapa 0,90 mm pestañada' if current_cat == 'Bandejas' else (' - chapa 0,71 mm' if current_cat == 'Pizzeras' else '')),
                    Precio=p_enlozada,
                    Moneda='ARS',
                    IVA=None,
                    **{'Archivo origen': 'Bandejas LGR.pdf', 'Pagina PDF': page_n},
                )
            if p_chapa and p_chapa > 100:
                yield row(
                    'LGR',
                    Categoria=cat,
                    Producto=f'{current_cat[:-1] if current_cat.endswith("s") else current_cat} chapa {measure}',
                    Descripcion='Chapa' + (' - 0,90 mm pestañada' if current_cat == 'Bandejas' else (' - 0,71 mm' if current_cat == 'Pizzeras' else '')),
                    Precio=p_chapa,
                    Moneda='ARS',
                    IVA=None,
                    **{'Archivo origen': 'Bandejas LGR.pdf', 'Pagina PDF': page_n},
                )


# ============ Don Segura Bateas ============
def parse_donsegura_bateas() -> Iterator[dict]:
    txt_path = TXT / 'Don Segura__LISTA BATEAS 01-05-2026.txt'
    if not txt_path.exists():
        return
    text = txt_path.read_text(encoding='utf-8', errors='replace')
    current_cat = 'Bateas amasadoras'
    for ln in text.splitlines():
        l = ln.strip()
        if not l:
            continue
        if l in ('AMASADORAS',):
            current_cat = 'Bateas amasadoras'
            continue
        # 'X LITROS ... $ 45,000 NOTAS' or 'TACHO BATIDORA ... $ 47,000'
        m = re.match(r'^(.+?)\s+\$\s*([\d.,]+)(?:\s+(.+?))?\s*$', l)
        if not m:
            continue
        producto = m.group(1).strip()
        price = to_money(m.group(2))
        notas = (m.group(3) or '').strip()
        if not price or price < 5000:
            continue
        # Skip lines that have leftover comma-prices like "3,90 U$S EL KG"
        if 'U$S' in l or 'I.V.A' in l:
            continue
        yield row(
            'Don Segura',
            Categoria=current_cat if 'TACHO' not in producto else 'Tachos batidora',
            Producto=producto,
            Descripcion=notas or None,
            Precio=price,
            Moneda='ARS',
            IVA='Sin IVA (+21%)',
            **{'Archivo origen': 'LISTA BATEAS 01-05-2026.pdf', 'Pagina PDF': 1},
        )


# ============ Asadores 2026 (image-only PDF, OCR) ============
def parse_asadores_ocr() -> Iterator[dict]:
    """Parse OCR of 'Asadores 2026.pdf' (image-only catalog).

    Strategy: the OCR is noisy and prices are rendered with stylized fonts that
    Tesseract can't read. We extract product NAMES (uppercase strings like
    'PARRILLERO BALCONERO') and emit them with price=None (Consultar precio).
    """
    txt_path = TXT_OCR / 'Asadores_2026.txt'
    if not txt_path.exists():
        return
    paired = lines_with_pages(txt_path)
    keyword_re = re.compile(r'^(PARRILLERO|ASADOR|BRASERO|DISCO|FOGON|FOGÓN|SALAMANDRA|TABLA)\b[A-ZÁÉÍÓÚÑ\s]*$')
    seen: set[str] = set()
    for pp, raw in paired:
        ln = raw.strip()
        if not ln or len(ln) > 60 or len(ln) < 6:
            continue
        if 'ASADORES' == ln or 'ACCESORIOS' in ln or 'ASADORESDEC' in ln:
            continue
        m = keyword_re.match(ln)
        if not m:
            continue
        name = re.sub(r'^[Y\s]+', '', ln).strip()
        if name in seen:
            continue
        seen.add(name)
        category = 'Parrilleros' if name.startswith('PARRILLERO') else 'Asadores'
        yield row(
            'Asadores DEC',
            Categoria=category,
            Producto=name.title(),
            Precio=None,
            Moneda=None,
            IVA=None,
            Observaciones='Catálogo Asadores 2026 — precio: consultar al proveedor',
            **{'Archivo origen': 'Asadores 2026.pdf', 'Pagina PDF': pp},
        )


# ============ MAIN ============
PARSERS = [
    ('Dunia XLSX', parse_dunia_xlsx),
    ('Havard Tramontina', parse_havard_tramontina),
    ('Havard Borgen', parse_havard_borgen),
    ('Havard Mayorista', lambda: parse_havard_catalog_text(
        TXT / 'Havard__HavardMayorista MAYO 2026.txt',
        'HavardMayorista MAYO 2026.pdf', 'Mayorista Skymsen/Varios')),
    ('Havard FAME (PDF)', parse_havard_fame_pdf),
    ('Havard Gelopar (PDF)', parse_havard_gelopar_pdf),
    ('Danda', parse_danda),
    ('Resinet (PDF)', parse_resinet_pdf),
    ('Fadeco', parse_fadeco),
    ('Don Segura Amasadoras', parse_donsegura_amasadoras),
    ('Don Segura Mesas y Piletas', parse_donsegura_mesas),
    ('Don Segura Bateas', parse_donsegura_bateas),
    ('DS Campanas', parse_ds_campanas),
    # MB Catálogo: el PDF es de marketing (productos en columnas con precios al pie
    # de página, separados visualmente). El extractor produce datos no confiables —
    # mejor flag para revisión manual hasta tener un layout estable.
    # ('MB Catalogo', parse_mb_catalogo),
    ('Turboblender DF Pricelist (OCR)', parse_turboblender_df_pricelist),
    ('Turboblender Línea TB (OCR)', parse_turboblender_lista_abril),
    ('Havard Preventa Börgen (OCR)', parse_havard_preventa_borgen),
    ('Bianchi NOVA 330 (OCR)', parse_bianchi_nova330),
    # New (Mayo 2026)
    ('ARE Precios', parse_are_precios),
    ('Picadoras y sierras', parse_picadoras_sierras),
    ('Calefactores de exterior', parse_calefactores_exterior),
    ('Mantenedores de calor', parse_mantenedores_calor),
    ('Bandejas LGR', parse_bandejas_lgr),
    ('Asadores 2026 (OCR)', parse_asadores_ocr),
]

NEEDS_REVIEW = [
    ('MB', 'CATALOGO DE PRODUCTOS MB.pdf', 'Catálogo de marketing con productos en columnas y precios al pie de página. La extracción automática produce datos poco confiables (ej. "ALERT!", "1 Velocidad", "Batea Acero Inox" con precios cruzados). Cargar manualmente o pedir lista de precios en formato tabular.'),
    ('Dunia', 'CATALOGO_MAQUINARIA_DUNIA_ABRIL_2026.pdf', 'Catálogo descriptivo (especificaciones técnicas) sin precios — usar XLSX consolidado para precios.'),
    ('Havard', 'FAM CATALOGO DIGITAL - HAVARD.pdf', 'Catálogo de especificaciones técnicas FAM sin precios — los precios FAME están en FAME HAVARD MARZO 2026.pdf.'),
    ('Don Segura', 'Don Segura.pdf', 'Catálogo descriptivo histórico sin precios — los precios viven en AMASADORAS Y SOBADORAS MAYORISTA 01-05-26.pdf y LISTA BATEAS 01-05-2026.pdf.'),
    ('Turboblender', 'CATALOGO DF V.02.25.pdf', 'Catálogo sin precios en el PDF — los precios están en LISTA DE PRECIOS DF V.01.26.pdf (ya extraído por OCR).'),
    ('Turboblender', 'Condición de pago Diciembre 2025.pdf', 'Condiciones de pago, no productos.'),
    ('DS Campanas', 'Nueva linea cristal.jpeg / semi circular.jpeg / Piramidal.jpeg / slim flat.jpeg', 'OCR aplicado: imágenes de catálogo con modelos pero sin precios legibles — usar precios de LISTA DE PRECIOS DS CAMPANAS.pdf.'),
    ('Havard', 'HavardMayorista-marzo-sinprecios.pdf', 'Catálogo sin precios (versión "sinprecios"). Los precios están en HavardMayorista MAYO 2026.pdf.'),
    ('Havard', 'PREVENTA CALEFACCION BORGEN 2026 -.pdf', 'OCR parcial: 2 de 3 productos extraídos; el segundo (BORGEN CON2000F) tiene precio ilegible — completar manualmente.'),
    ('Turboblender', 'Lista Turboblender abril 26.pdf', 'OCR parcial: precios renderizados con tipografía estilizada y mayormente ilegibles por OCR — completar manualmente.'),
    ('ARE', 'Catalogo ARE - Diciembre 2024.pdf', 'Catálogo descriptivo (especificaciones técnicas) sin precios — los precios están en PRECIOS.pdf.'),
    ('Asadores DEC', 'Asadores 2026.pdf', 'Catálogo de imágenes (12 páginas) — OCR best-effort; la mayoría de los productos quedan sin precio. Consultar al proveedor.'),
]


def main():
    all_rows = []
    per_supplier: dict[str, list] = {}
    errors = []
    for label, fn in PARSERS:
        try:
            rows = list(fn())
        except Exception as e:
            errors.append((label, str(e)))
            traceback.print_exc()
            continue
        print(f'  [{label}] -> {len(rows)} items')
        all_rows.extend(rows)
    print(f'TOTAL rows (before dedup): {len(all_rows)}')

    # ===== Deduplication pass =====
    # Key: (proveedor, producto-normalizado, precio). Keeps the first occurrence.
    def norm(s):
        if not s:
            return ''
        return re.sub(r'\s+', ' ', str(s)).strip().lower()

    seen_keys: dict[tuple, dict] = {}
    deduped_rows = []
    dup_count = 0
    for r in all_rows:
        key = (r.get('Proveedor'), norm(r.get('Producto')), r.get('Precio'), r.get('Moneda'))
        if key in seen_keys:
            dup_count += 1
            continue
        seen_keys[key] = r
        deduped_rows.append(r)
    all_rows = deduped_rows
    if dup_count:
        print(f'  Deduped: {dup_count} filas eliminadas')

    # ===== Validation pass =====
    # Heuristics: ARS price < 100 is suspicious; same code across two prices is suspicious.
    warnings_list = []
    by_supplier_code: dict[tuple, list] = {}
    for r in all_rows:
        if r.get('Precio') is not None:
            try:
                p = float(r['Precio'])
            except (TypeError, ValueError):
                p = None
            if p is not None:
                if r.get('Moneda') == 'ARS' and p < 100:
                    warnings_list.append((r['Proveedor'], r.get('Codigo'), r.get('Producto'), f'precio ARS sospechosamente bajo: {p}'))
                if r.get('Moneda') == 'USD' and p > 50000:
                    warnings_list.append((r['Proveedor'], r.get('Codigo'), r.get('Producto'), f'precio USD sospechosamente alto: {p}'))
        if r.get('Codigo'):
            by_supplier_code.setdefault((r['Proveedor'], r['Codigo']), []).append(r)

    # Same supplier+code with multiple different prices → suspect duplicate
    for (sup, code), items in by_supplier_code.items():
        prices = {(it.get('Precio'), it.get('Moneda')) for it in items if it.get('Precio') is not None}
        if len(prices) > 1:
            warnings_list.append((sup, code, items[0].get('Producto'), f'mismo código con {len(prices)} precios distintos'))

    if warnings_list:
        print(f'  Validation warnings: {len(warnings_list)}')
        for w in warnings_list[:10]:
            print(f'    [{w[0]}] {w[1] or "(sin código)"} | {(w[2] or "")[:40]} → {w[3]}')
        if len(warnings_list) > 10:
            print(f'    ... y {len(warnings_list) - 10} más')

    # Rebuild per-supplier
    per_supplier.clear()
    for r in all_rows:
        per_supplier.setdefault(r['Proveedor'], []).append(r)
    print(f'TOTAL rows (final): {len(all_rows)} en {len(per_supplier)} proveedores')

    # Write Excel
    df_all = pd.DataFrame(all_rows, columns=COLUMNS)
    df_review = pd.DataFrame(NEEDS_REVIEW, columns=['Proveedor', 'Archivo', 'Motivo'])

    with pd.ExcelWriter(OUT, engine='xlsxwriter') as w:
        df_all.to_excel(w, sheet_name='Consolidado', index=False)
        # Per-supplier sheets (Excel sheet names max 31 chars, no special chars)
        seen = set()
        for supplier, items in per_supplier.items():
            name = re.sub(r'[\[\]\*\?/\\:]', '', supplier)[:31]
            if name in seen:
                name = name[:28] + '_2'
            seen.add(name)
            pd.DataFrame(items, columns=COLUMNS).to_excel(w, sheet_name=name, index=False)
        df_review.to_excel(w, sheet_name='Para revisar', index=False)
        if errors:
            pd.DataFrame(errors, columns=['Parser', 'Error']).to_excel(w, sheet_name='Errores', index=False)
        if warnings_list:
            pd.DataFrame(warnings_list, columns=['Proveedor', 'Código', 'Producto', 'Warning']).to_excel(
                w, sheet_name='Validación', index=False
            )

        # Auto-format columns
        for sheetname in w.sheets:
            ws = w.sheets[sheetname]
            ws.set_column(0, 0, 14)  # Proveedor
            ws.set_column(1, 1, 28)  # Categoria
            ws.set_column(2, 2, 18)  # Codigo
            ws.set_column(3, 3, 50)  # Producto
            ws.set_column(4, 4, 40)  # Descripcion
            ws.set_column(5, 5, 14)  # Precio
            ws.set_column(6, 6, 8)   # Moneda
            ws.set_column(7, 7, 10)  # IVA
            ws.set_column(8, 8, 12)  # Cantidad x Bulto
            ws.set_column(9, 9, 18)  # Codigo Barras
            ws.set_column(10, 10, 28) # Observaciones
            ws.set_column(11, 11, 36) # Archivo origen

    print(f'Wrote {OUT}')


if __name__ == '__main__':
    main()
